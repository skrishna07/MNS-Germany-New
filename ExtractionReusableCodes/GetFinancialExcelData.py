import openpyxl
import pandas as pd
from pathlib import Path


def extract_all_sheets(filepath):
    wb = openpyxl.load_workbook(filepath)
    all_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            # Keep rows that have at least one non-None value
            if any(cell is not None for cell in row):
                rows.append(list(row))

        if rows:
            # Use first row as header, replace None/NaN with empty string
            headers = ["" if h is None else h for h in rows[0]]
            data_rows = rows[1:]
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.fillna("").astype(str).replace("None", "")
            all_data[sheet_name] = df

    return all_data


def display_all_data(all_data):
    extracted_data = ""
    for sheet_name, df in all_data.items():
        # print(f"\n{'=' * 60}")
        # print(f"  Sheet: {sheet_name}  ({len(df)} rows x {len(df.columns)} columns)")
        # print(f"{'=' * 60}")
        # print(df.to_string(index=False, na_rep=""))
        extracted_data += f"\n{'=' * 60}\n"
        extracted_data += f"  Sheet: {sheet_name}  ({len(df)} rows x {len(df.columns)} columns)\n"
        extracted_data += f"{'=' * 60}\n"
        extracted_data += df.to_string(index=False, na_rep="") + "\n"
    return extracted_data


def save_to_excel(all_data, output_path="extracted_data.xlsx"):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in all_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, na_rep="")
    print(f"\nAll sheets saved to: {output_path}")


def save_to_csv(all_data, output_dir="extracted_csv"):
    Path(output_dir).mkdir(exist_ok=True)
    for sheet_name, df in all_data.items():
        csv_path = f"{output_dir}/{sheet_name}.csv"
        df.to_csv(csv_path, index=False)
        print(f"  Saved: {csv_path}")


def get_excel_data(input_excel_path):
    try:
        all_data = extract_all_sheets(input_excel_path)

        # 1. Print all tables to console
        extracted_data = display_all_data(all_data)
        print(extracted_data)
        return extracted_data
    except Exception as e:
        raise Exception(f"Error in getting data from excel {e}")

# if __name__ == "__main__":
#     all_data = extract_all_sheets(FILE_PATH)
#
#     # 1. Print all tables to console
#     extracted_data = display_all_data(all_data)
#     print(extracted_data)
#
#     # 2. Save all sheets into one Excel file
#     # save_to_excel(all_data, OUTPUT_EXCEL)
#
#     # 3. (Optional) Save each sheet as a separate CSV
#     # save_to_csv(all_data, OUTPUT_CSV_DIR)